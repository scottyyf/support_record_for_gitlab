#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
File: ExcelActor.py.py
Author: Scott Yang(Scott)
Email: yangyingfa@skybility.com
Copyright: Copyright (c) 2021, Skybility Software Co.,Ltd. All rights reserved.
Description:
"""
import os
import logging
import re
import shutil
from abc import abstractmethod

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet import worksheet

import DataCollect


def int_to_upper(int_value):
    return chr(int_value + 64)


class SheetHandler:
    # SHEET_TITLE = {title1: width1,...}
    SHEET_TITLE = {}

    def __init__(self, sheet: worksheet):
        self.sheet = sheet
        self.data = []
        self.add_title_if_empty()

    @abstractmethod
    def append_a_report(self, a_reporter: dict):
        pass

    def add_title_if_empty(self):
        mark = True
        for i in range(1, 4):
            if self.sheet[f'{int_to_upper(i)}1'].value is not None:
                mark = False
                break

        if mark:
            self.add_title_to_head()

    def add_title_to_head(self):
        col = 1
        for value, width in self.SHEET_TITLE.items():
            # col = int_to_upper(col)
            self.add_to_cell(f'{int_to_upper(col)}1', value)
            self.set_col_width(f'{int_to_upper(col)}', width)
            col += 1

        self.set_row_height(1, 40)

    def add_to_cell(self, cell, value):
        self.sheet[cell] = value
        col, row = coordinate_from_string(cell)
        if col == 'A':
            self.set_row_height(row, 40)

        self.cell_wrap_text(cell)

    def cell_wrap_text(self, cell):
        self.sheet[cell].alignment = Alignment(wrapText=True,
                                               horizontal='left',
                                               vertical='center')

    def set_col_width(self, col, width):
        self.sheet.column_dimensions[col].width = width

    def set_row_height(self, row, height):
        self.sheet.row_dimensions[row].height = height

    def get_cell_data(self, row, col):
        if isinstance(col, int):
            col = int_to_upper(col)

        cell = f'{col}{row}'
        return self.sheet[cell].value

    def set_zone_border(self):
        # return
        rws = self.sheet.max_row
        cls = self.sheet.max_column
        thin = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

        for cl in range(1, cls + 1):
            for rw in range(1, rws + 1):
                self.sheet.cell(row=rw, column=cl).border = thin


class Ha28DetailSheet(SheetHandler):
    SHEET_TITLE = {
        '问题类型': 9.8,
        '问题提交人': 23.55,
        '时间': 10.93,
        '联系电话': 10,
        '联系邮箱': 29.3,
        '现场': 9.93,
        '处理工程师': 11.93,
        '版本': 9.93,
        '邮箱主题': 31.68,
        '问题描述': 9.93,
        '问题分析': 30.68,
        '方案建议': 29.93,
        }

    def append_a_report(self, a_reporter: dict):
        row = self.sheet.max_row + 1
        self.add_to_cell(f'A{row}', 'Ha2.8')
        self.add_to_cell(f'B{row}', a_reporter.get('who'))
        self.add_to_cell(f'C{row}', a_reporter.get('answer_at'))
        self.add_to_cell(f'D{row}', '')
        self.add_to_cell(f'E{row}', a_reporter.get('email'))
        self.add_to_cell(f'F{row}', '')
        self.add_to_cell(f'G{row}', a_reporter.get('our_people'))
        self.add_to_cell(f'H{row}', '')
        self.add_to_cell(f'I{row}', a_reporter.get('title'))
        self.add_to_cell(f'J{row}', a_reporter.get('title'))
        self.add_to_cell(f'K{row}', a_reporter.get('fenxi'))
        self.add_to_cell(f'L{row}', a_reporter.get('jianyi'))


class Ha28CustomerSheet(SheetHandler):
    SHEET_TITLE = {
        '问题提交人': 29.3,
        '联系电话': 29.3,
        '联系邮箱': 29.3,
        }

    def append_a_report(self, a_reporter: dict):
        if self._check_dump(a_reporter):
            return

        row = self.sheet.max_row + 1
        self.add_to_cell(f'A{row}', a_reporter.get('who'))

        # logging.info(row)
        if re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
                     a_reporter.get('email')):  # email
            self.add_to_cell(f'C{row}', a_reporter.get('email'))
        else:
            self.add_to_cell(f'B{row}', a_reporter.get('email'))

    def _check_dump(self, a_reporter):
        all_contact = set()
        # logging.info(self.sheet.max_row + 1)
        for row in range(2, self.sheet.max_row + 1):
            all_contact.add(self.get_cell_data(row, 1))

        contact = a_reporter.get('who')
        if contact in all_contact:
            return True

        return False


# 用户自定义地方
class UserWanted(SheetHandler):
    SHEET_TITLE = {}

    def append_a_report(self, a_reporter: dict):
        pass


class XlsHandler:
    def __init__(self, file_name):
        self.file_name = file_name
        self._work_book = Workbook()
        self.load_workbook()
        self._sheets = {}

    def load_workbook(self):
        if os.path.isfile(self.file_name):
            try:
                self._work_book = load_workbook(self.file_name)
            except Exception as e:
                logging.warning(
                    f'Failed to load file {self.file_name}, detail: {e}')

    def load_sheet_if_exist(self, sheet_name='', insert_index=None):
        if not sheet_name:
            return

        if sheet_name in self._work_book.sheetnames:
            self._sheets[sheet_name] = self._work_book[sheet_name]
            return

        work_sheet = self._work_book.create_sheet(index=insert_index,
                                                  title=sheet_name)
        self._sheets[sheet_name] = work_sheet

    @abstractmethod
    def add_new_data(self, data):
        pass

    def back_old_file(self):
        if os.path.isfile(self.file_name):
            tmp_file = self.file_name.replace('xlsx', 'bak.xlsx')
            logging.info(f'Backup xls file {self.file_name} to {tmp_file}')
            shutil.copy2(self.file_name, tmp_file)

    def write_new_file(self):
        logging.info(f'Write new data to {self.file_name}')
        self._work_book.save(self.file_name)


class Ha28xlsHandler(XlsHandler):
    sheet1_name = 'HA服务记录'
    sheet2_name = '用户记录'

    def add_new_data(self, data):
        self.load_sheet_if_exist(self.sheet1_name, 0)
        self.load_sheet_if_exist(self.sheet2_name, 1)

        detail_sh = Ha28DetailSheet(self._sheets[self.sheet1_name])
        customer_sh = Ha28CustomerSheet(self._sheets[self.sheet2_name])
        for _data in data:
            detail_sh.append_a_report(_data)
            customer_sh.append_a_report(_data)

        detail_sh.set_zone_border()
        customer_sh.set_zone_border()


# 用户自定义表名
class UserWantxlHandler(XlsHandler):
    sheet1_name = ''
    sheet2_name = ''

    def add_new_data(self, data):
        pass


class ExcelActor:
    # 用户增加excel对应的xlshandler
    ONE_TO_ONE = {'ha2.8': Ha28xlsHandler}

    def __init__(self, file_name='', data: DataCollect = None, family='suse'):
        self.data = data
        self.family = self.ONE_TO_ONE.get(family, Ha28xlsHandler)(file_name)
        # self.do()

    def do(self):
        # self.load_current_data()
        self.family.add_new_data(self.data)
        self.family.back_old_file()
        self.family.write_new_file()
